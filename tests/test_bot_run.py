from urllib.parse import unquote_plus
import httpx
from content_factory.publish.telegram import PublishState, PublishResult
from content_factory.orchestrator.confirm_store import Awaiting
from content_factory.bot import run as botrun


def test_get_updates_parses_result():
    def handler(req):
        assert req.url.path == "/botTOK/getUpdates"
        return httpx.Response(200, json={"ok": True, "result": [{"update_id": 5}]})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    upd = botrun.get_updates("TOK", 0, timeout=0, http=http)
    assert upd == [{"update_id": 5}]


def test_finalize_preview_replaces_buttons_with_verdict():
    reqs = []

    def handler(req):
        reqs.append((req.url.path, req.read()))
        return httpx.Response(200, json={"ok": True})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    cq = {"message": {"chat": {"id": -100777}, "message_id": 42}}
    botrun.finalize_preview(http, "TOK", cq, "✅ опубликовано: k1")
    path, body = reqs[0]
    assert path == "/botTOK/editMessageReplyMarkup"
    decoded = unquote_plus(body.decode())          # form-URL-encoded → текст
    assert "message_id=42" in decoded and "✅ опубликовано: k1" in decoded


def test_finalize_preview_no_message_is_noop():
    def handler(req):
        raise AssertionError("не должно быть запросов")
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    botrun.finalize_preview(http, "TOK", {}, "✅")   # без message — тихо выходим


def test_make_publish_fn_uses_awaiting_channel(tmp_path):
    captured = {}

    def handler(req):
        captured["path"] = req.url.path
        return httpx.Response(200, json={"ok": True, "result": {"message_id": 7}})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    ps = PublishState(tmp_path / "p.db")
    fn = botrun.make_publish_fn("TOK", "HTML", ps, http=http)
    res = fn(Awaiting(key="k1", channel="@chan", card_path="https://x/c.jpg",
                      caption="cap", status="pending"))
    assert res.ok and res.message_id == 7
    assert captured["path"] == "/botTOK/sendPhoto"


def test_make_regen_fn_removes_card_and_store_entry(tmp_path):
    import sqlite3
    card = tmp_path / "NC_123.jpg"
    card.write_bytes(b"IMG")
    store_db = tmp_path / "cards.db"
    with sqlite3.connect(store_db) as c:
        c.execute("CREATE TABLE card_jobs (key TEXT PRIMARY KEY, input_filename TEXT, "
                  "status TEXT, tries INTEGER DEFAULT 0)")
        c.execute("INSERT INTO card_jobs VALUES ('NC_123', 'in.jpg', 'done', 1)")
    fn = botrun.make_regen_fn(store_db, tmp_path / "state.db")
    a = Awaiting(key="breeze|x|y", channel="@c", card_path=str(card),
                 caption="cap", status="published")
    assert fn(a) is True
    assert not card.exists()                       # файл карточки удалён
    with sqlite3.connect(store_db) as c:
        assert c.execute("SELECT count(*) FROM card_jobs WHERE key='NC_123'").fetchone()[0] == 0


def test_make_regen_fn_survives_missing_file(tmp_path):
    import sqlite3
    store_db = tmp_path / "cards.db"
    with sqlite3.connect(store_db) as c:
        c.execute("CREATE TABLE card_jobs (key TEXT PRIMARY KEY, input_filename TEXT, "
                  "status TEXT, tries INTEGER DEFAULT 0)")
    fn = botrun.make_regen_fn(store_db, tmp_path / "state.db")
    a = Awaiting(key="k", channel="@c", card_path=str(tmp_path / "нет_файла.jpg"),
                 caption="cap", status="pending")
    assert fn(a) is True                           # отсутствие файла/записи — не ошибка


def test_make_regen_fn_resets_excel_item_to_new(tmp_path):
    # Грабля 2026-07-06 (владелец: «отпускаю на перегенерацию — не перегенерируется»):
    # для excel-товара удаление карточки недостаточно — excel_items оставался в
    # preview, а тик пересобирает только new/research/card. Теперь regen возвращает
    # товар в new: research возьмётся из кэша → сразу новая карточка.
    from content_factory.orchestrator.excel_pipeline import ExcelStore
    import sqlite3
    store_db = tmp_path / "cards.db"
    with sqlite3.connect(store_db) as c:
        c.execute("CREATE TABLE card_jobs (key TEXT PRIMARY KEY, input_filename TEXT, "
                  "status TEXT, tries INTEGER DEFAULT 0)")
    state_db = tmp_path / "state.db"
    es = ExcelStore(state_db)
    es.add_items([("excel|lg|ga-b419slgl", "LG", "GA-B419SLGL",
                   "Холодильник LG GA-B419SLGL", 45000)])
    es.update("excel|lg|ga-b419slgl", status="preview", research_job=568, card_job=597)

    card = tmp_path / "excel_lg-ga-b419slgl.jpg"
    card.write_bytes(b"IMG")
    fn = botrun.make_regen_fn(store_db, state_db)
    a = Awaiting(key="excel|lg|ga-b419slgl", channel="@c", card_path=str(card),
                 caption="cap", status="published")
    assert fn(a) is True
    assert not card.exists()
    item = [i for i in es.by_status("new") if i.key == "excel|lg|ga-b419slgl"]
    assert len(item) == 1                          # товар вернулся в конвейер
    assert item[0].research_job is None and item[0].card_job is None
    assert item[0].tries == 0


def test_publish_fn_adds_order_button(tmp_path):
    from content_factory.publish.orders import OrderLinks
    captured = {}

    def handler(req):
        captured["body"] = req.read()
        return httpx.Response(200, json={"ok": True, "result": {"message_id": 7}})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    ps = PublishState(tmp_path / "p.db")
    links = OrderLinks(tmp_path / "p.db")
    fn = botrun.make_publish_fn("TOK", "HTML", ps, http=http,
                                order_bot="Sendpr1ce_bot", links=links)
    res = fn(Awaiting(key="k1", channel="@chan", card_path="https://x/c.jpg",
                      caption="cap", status="pending"))
    assert res.ok
    decoded = unquote_plus(captured["body"].decode())
    assert "t.me/Sendpr1ce_bot?start=ord_" in decoded      # кнопка «Заказать» в посте


def test_make_fn_selects_and_stores(tmp_path):
    import openpyxl
    from content_factory.orchestrator.excel_pipeline import ExcelStore
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    ws.append(["Холодильники", "", "", "", "", ""])
    ws.append(["1", "10", "Beko", "Холодильник Beko X100", "30000", ""])
    ws.append(["2", "11", "Candy", "Холодильник Candy Y200", "25000", ""])
    prices = tmp_path / "prices"
    prices.mkdir()
    wb.save(prices / "manual.xlsx")
    fn = botrun.make_make_fn(tmp_path / "state.db", prices)
    reply = fn(2, "холодильники", {})
    assert "✅ выбрано 2" in reply and "Beko X100" in reply
    items = ExcelStore(tmp_path / "state.db").by_status("new")
    assert {i.key for i in items} == {"excel|beko|x100", "excel|candy|y200"}
    reply2 = fn(2, "холодильники", {})                    # повтор — всё уже в работе
    assert "❌" in reply2


def test_make_fn_searches_both_manual_and_mail_slots(tmp_path):
    # грабля 2026-07-03: почта (cf-mail) молча затирала единственный latest.xlsx —
    # позиции владельца из его собственного прайса переставали находиться.
    import openpyxl
    from content_factory.orchestrator.excel_pipeline import ExcelStore
    prices = tmp_path / "prices"
    prices.mkdir()

    wb1 = openpyxl.Workbook(); ws1 = wb1.active
    ws1.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    ws1.append(["Холодильники", "", "", "", "", ""])
    ws1.append(["1", "10", "Beko", "Холодильник Beko X100", "30000", ""])
    wb1.save(prices / "manual.xlsx")

    wb2 = openpyxl.Workbook(); ws2 = wb2.active
    ws2.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    ws2.append(["Холодильники", "", "", "", "", ""])
    ws2.append(["2", "11", "Candy", "Холодильник Candy Y200", "25000", ""])
    wb2.save(prices / "mail.xlsx")

    fn = botrun.make_make_fn(tmp_path / "state.db", prices)
    reply = fn(2, "холодильники", {})
    assert "Beko X100" in reply and "Candy Y200" in reply     # найдено в обоих слотах
    items = ExcelStore(tmp_path / "state.db").by_status("new")
    assert {i.key for i in items} == {"excel|beko|x100", "excel|candy|y200"}


# ── download_telegram_file: общий хелпер getFile → байты (визард /task, шаг 4) ──
def test_download_telegram_file_success():
    def handler(req):
        if req.url.path == "/botTOK/getFile":
            return httpx.Response(200, json={"ok": True,
                                             "result": {"file_path": "photos/file_1.jpg"}})
        if req.url.path == "/file/botTOK/photos/file_1.jpg":
            return httpx.Response(200, content=b"IMGBYTES")
        raise AssertionError(f"неожиданный путь: {req.url.path}")
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    data = botrun.download_telegram_file(http, "TOK", "fid123")
    assert data == b"IMGBYTES"


def test_download_telegram_file_missing_path_returns_none():
    def handler(req):
        return httpx.Response(200, json={"ok": True, "result": {}})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    assert botrun.download_telegram_file(http, "TOK", "fid123") is None


def test_receive_price_saves_to_manual_slot(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    ws.append(["Холодильники", "", "", "", "", ""])
    ws.append(["1", "10", "Beko", "Холодильник Beko X100", "30000", ""])
    import io
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    def handler(req):
        if req.url.path == "/botTOK/getFile":
            return httpx.Response(200, json={"ok": True,
                                             "result": {"file_path": "documents/f.xlsx"}})
        if req.url.path == "/file/botTOK/documents/f.xlsx":
            return httpx.Response(200, content=xlsx_bytes)
        raise AssertionError(f"неожиданный путь: {req.url.path}")
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    reply = botrun.receive_price(http, "TOK", {"file_id": "fid", "file_name": "price.xlsx"},
                                 tmp_path)

    from content_factory.ingest.excel_price import manual_slot_name
    assert "1 позиций" in reply and "Холодильники" in reply
    # ручной прайс — в свой слот поставщика (не общий manual.xlsx)
    assert (tmp_path / f"{manual_slot_name('price.xlsx')}.xlsx").read_bytes() == xlsx_bytes
    assert (tmp_path / "price.xlsx").read_bytes() == xlsx_bytes


def test_receive_price_download_failure():
    def handler(req):
        return httpx.Response(200, json={"ok": True, "result": {}})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    reply = botrun.receive_price(http, "TOK", {"file_id": "fid"}, "unused")
    assert "❌" in reply


def test_receive_price_saves_to_named_slot(tmp_path):
    # источник-канал (2026-07-04): та же логика, другой слот (не перезаписывает
    # manual.xlsx владельца — load_price_slots ищет во всех трёх)
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    ws.append(["Телевизоры", "", "", "", "", ""])
    ws.append(["1", "20", "MIU", "Телевизор MIU H40", "20000", ""])
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    def handler(req):
        if req.url.path == "/botTOK/getFile":
            return httpx.Response(200, json={"ok": True,
                                             "result": {"file_path": "documents/f.xlsx"}})
        if req.url.path == "/file/botTOK/documents/f.xlsx":
            return httpx.Response(200, content=xlsx_bytes)
        raise AssertionError(f"неожиданный путь: {req.url.path}")
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    reply = botrun.receive_price(http, "TOK", {"file_id": "fid", "file_name": "БытТехОпт.xlsx"},
                                 tmp_path, slot="channel")

    assert "1 позиций" in reply
    assert (tmp_path / "channel.xlsx").read_bytes() == xlsx_bytes
    assert not (tmp_path / "manual.xlsx").exists()   # свой прайс владельца не задет


# ── /excel: читаемый статус (жалоба владельца 2026-07-07: «всё слитно») ──────
def test_excel_fn_truncates_multiline_errors_and_separates_sections(tmp_path):
    from content_factory.orchestrator.excel_pipeline import ExcelStore
    es = ExcelStore(tmp_path / "state.db")
    es.add_items([("excel|don|r-103", "DON", "R-103", "Морозильная камера DON R-103", 9000),
                  ("excel|lg|x", "LG", "X", "Холодильник LG X", 30000)])
    es.update("excel|don|r-103", status="failed",
              error="research: Page.wait_for_selector: Timeout 20000ms exceeded.\n"
                    "Call log:\n  - waiting for locator(\"div\") to be visible\n)")
    es.update("excel|lg|x", status="card", card_job=1)

    _, _, excel_fn = botrun.make_find_pick_fns(tmp_path / "state.db", tmp_path)
    text = excel_fn()

    assert "Call log" not in text                  # простыня ошибки обрезана
    assert "Timeout 20000ms" in text               # суть ошибки осталась
    assert "─" in text or "—" in text or "━" in text   # есть разделители секций


def test_resolve_callback_data_expands_code(tmp_path):
    from content_factory.publish.orders import OrderLinks
    from content_factory.orchestrator.confirm_store import ConfirmStore
    cs = ConfirmStore(tmp_path / "s.db")
    links = OrderLinks(tmp_path / "s.db")
    long_key = "excel|генератор бензиновый carver ppg - 1200i cube инверторный"
    cs.add(long_key, "@chan", "/c/x.jpg", "cap")
    code = links.code_for(long_key)
    assert botrun.resolve_callback_data(f"approve:{code}", cs, links) == f"approve:{long_key}"
    # обычные короткие ключи проходят как есть
    cs.add("breeze|funai|daijin", "@chan", "/c/y.jpg", "cap")
    assert botrun.resolve_callback_data("approve:breeze|funai|daijin", cs, links) == \
        "approve:breeze|funai|daijin"
    assert botrun.resolve_callback_data("noop", cs, links) == "noop"


def test_setup_bot_commands_includes_auto():
    # /auto должен попасть в меню владельца (setMyCommands применяется при старте бота)
    reqs = []

    def handler(req):
        reqs.append((req.url.path, unquote_plus(req.read().decode())))
        return httpx.Response(200, json={"ok": True})
    http = httpx.Client(transport=httpx.MockTransport(handler), base_url="https://api.telegram.org")
    botrun.setup_bot_commands(http, "TOK", "42")
    assert len(reqs) == 2
    path, owner_body = reqs[0]
    assert path == "/botTOK/setMyCommands"
    assert '"command": "auto"' in owner_body           # выключатель автомата в меню
    assert '"command": "status"' in owner_body
    _, default_body = reqs[1]
    assert "commands=[]" in default_body               # у клиентов меню пустое


def test_auto_markup_toggle_button():
    # владелец просил КНОПКИ вкл/выкл (2026-07-09), не текстовые команды
    on = botrun.auto_markup(True)["inline_keyboard"][0][0]
    assert on["callback_data"] == "auto:off" and "Выключить" in on["text"]
    off = botrun.auto_markup(False)["inline_keyboard"][0][0]
    assert off["callback_data"] == "auto:on" and "Включить" in off["text"]
    # п.5: полный контроль — кнопки времени/количества/категорий и сброса
    flat = str(botrun.auto_markup(True))
    for cb in ("auto:ask:times", "auto:ask:count", "auto:ask:cats", "auto:reset"):
        assert cb in flat


def test_markup_fn_db_sources_and_listing(tmp_path):
    # п.7 (2026-07-09): наценка на ВСЕ виды товара — БД-источники и '*' из бота
    fn = botrun.make_markup_fn(tmp_path / "prices", state_db=tmp_path / "s.db")
    assert "breeze" in fn("breeze", -3)
    assert "8" in fn("*", 8)
    from content_factory.pricing.overrides import markup_overrides
    assert markup_overrides(tmp_path / "s.db") == {"breeze": -3.0, "*": 8.0}
    out = fn("", None)                                     # обзор
    assert "breeze" in out and "-3" in out
    assert "❌" in fn("noexist", 5)                        # ни БД-источник, ни xlsx
    fn("breeze", 0)                                        # 0 = убрать (как в excel)
    assert markup_overrides(tmp_path / "s.db") == {"*": 8.0}
