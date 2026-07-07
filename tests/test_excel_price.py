"""Excel-прайс (БытТехОпт): парсер, извлечение модели, выбор позиций под /make."""
import openpyxl
from content_factory.ingest.excel_price import (
    parse_price_xlsx, extract_model, item_key, select_from_price)


def _xlsx(tmp_path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([""] * 6)                                    # пустая строка сверху (как в реале)
    ws.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    for r in rows:
        ws.append(r)
    p = tmp_path / "price.xlsx"
    wb.save(p)
    return p


ROWS = [
    ["Холодильники с нижней морозильной камерой", "", "", "", "", ""],
    ["1", "1210", "Gorenje", "Холодильник Gorenje NRK6201ES4", "40451", ""],
    ["2", "1211", "Beko", "Холодильник Beko B1RCSK362S", "25990", ""],
    ["Стиральные машины с фронтальной загрузкой", "", "", "", "", ""],
    ["3", "1300", "Beko", "Стиральная машина Beko WSRE6512", "21990", ""],
    ["4", "1301", "Candy", "Стиральная машина Candy CS4", "19990", ""],
    ["Stinol", "", "", "", "", ""],
    ["5", "1400", "Stinol", "Холодильник Stinol STS 167 (167*60*62)", "20254", ""],
]


def test_parse_price_sections_and_items(tmp_path):
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    assert len(items) == 5
    assert items[0].section == "Холодильники с нижней морозильной камерой"
    assert items[0].brand == "Gorenje" and items[0].price == 40451
    assert items[2].section == "Стиральные машины с фронтальной загрузкой"
    assert items[4].brand == "Stinol" and items[4].section == "Stinol"


def test_parse_skips_rows_without_price(tmp_path):
    rows = ROWS + [["6", "1500", "X", "Товар без цены", None, ""]]
    items = parse_price_xlsx(_xlsx(tmp_path, rows))
    assert all(i.price for i in items)


def test_extract_model_real_names():
    assert extract_model("Холодильник Stinol STS 167 (167*60*62)", "Stinol") == "STS 167"
    assert extract_model("Встраиваемый холодильник Beko BCNA306E2S ( А+ )", "Beko") == "BCNA306E2S"
    assert extract_model("Холодильник Gorenje NRK6201ES4", "Gorenje") == "NRK6201ES4"
    assert extract_model("Просто товар без бренда", "Nope") == "Просто товар без бренда"


def test_item_key_matches_manual_pilot_keys(tmp_path):
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    stinol = [i for i in items if i.brand == "Stinol"][0]
    assert item_key(stinol) == "excel|stinol|sts 167"      # совпадает с ключами пилота


def test_select_quotas_and_rest(tmp_path):
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    got = select_from_price(items, "холодильник", {"beko": 1, "*": None}, 3, taken=set())
    brands = [i.brand.lower() for i in got]
    assert len(got) == 3 and brands.count("beko") == 1     # квота Beko + добор остальными
    assert not any("Стиральная" in i.name for i in got)    # категория держится (не стиралки)


def test_select_respects_taken_keys(tmp_path):
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    taken = {"excel|stinol|sts 167"}
    got = select_from_price(items, "холодильник", {}, 10, taken=taken)
    assert all(item_key(i) != "excel|stinol|sts 167" for i in got)


def test_select_category_filters_sections(tmp_path):
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    got = select_from_price(items, "стиральные", {}, 10, taken=set())
    assert {i.brand for i in got} == {"Beko", "Candy"}


# ── формат «1С-карточки» (прайс ИП Аксёнов): блоки строк, без колонки бренда ──
def _xlsx_1c(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([""] * 14)
    row = [""] * 14
    row[1] = "1. ИНСТРУМЕНТ ДЛЯ СТРОЙКИ"
    ws.append(row)
    row = [""] * 14
    row[1] = "Генераторы"
    ws.append(row)
    def block(name, desc, code, art, price):
        r = [""] * 14
        r[1], r[5] = "НЕТ\nФОТОГРАФИИ", name
        ws.append(r)
        if desc:
            r = [""] * 14
            r[5] = desc
            ws.append(r)
        r = [""] * 14
        r[5], r[9], r[13] = "Код", "Артикул", "Цена"
        ws.append(r)
        r = [""] * 14
        r[5], r[9], r[13] = code, art, price
        ws.append(r)
        ws.append([""] * 14)
    block("Генератор бензиновый ВИТЯЗЬ БГ-8700", "Мощность 8.7 кВт\nБак 25 л",
          "УТ-00007786", "18037001", "37 520,00 RUB")
    block("Генератор инверторный Huter DN2700i", None,
          "УТ-00009354", "18037004", "28 140,00 RUB")
    p = tmp_path / "akse.xlsx"
    wb.save(p)
    return p


def test_parse_1c_blocks(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    assert len(items) == 2
    assert items[0].name == "Генератор бензиновый ВИТЯЗЬ БГ-8700"
    assert items[0].price == 37520
    assert items[0].section == "Генераторы"
    assert items[0].article == "18037001"
    assert items[1].price == 28140


def test_1c_key_without_brand_column(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    assert item_key(items[0]).startswith("excel|")
    assert "витязь" in item_key(items[0])


def test_select_1c_by_word_in_name(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    got = select_from_price(items, "генератор", {"huter": 1, "*": None}, 2, taken=set())
    assert len(got) == 2
    got1 = select_from_price(items, "генератор", {"huter": 1}, 1, taken=set())
    assert "Huter" in got1[0].name                          # квота по слову в имени


def test_select_prefers_name_match_over_section(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    # добавим в тот же раздел «Генераторы» товар БЕЗ слова в имени
    from content_factory.ingest.excel_price import PriceItem
    items.insert(0, PriceItem(section="Генераторы", article="1", brand="",
                              name="Автомат ввода резерва CARVER", price=10000))
    got = select_from_price(items, "генератор", {}, 2, taken=set())
    assert all("Генератор" in i.name for i in got)          # сами генераторы — первыми


# ── поиск без падежей/окончаний + фраза из нескольких слов ────────────────────
def test_stem_matching_cases(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    for phrase in ("генератора", "генераторы", "генераторов", "ГЕНЕРАТОР"):
        got = select_from_price(items, phrase, {}, 5, taken=set())
        assert len(got) == 2, phrase


def test_multiword_phrase_narrows(tmp_path):
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    got = select_from_price(items, "генераторы инверторные", {}, 5, taken=set())
    assert len(got) == 1 and "Huter" in got[0].name        # сузилось до инверторного


def test_search_items_numbered(tmp_path):
    from content_factory.ingest.excel_price import search_items
    items = parse_price_xlsx(_xlsx_1c(tmp_path))
    found = search_items(items, "генераторы", taken=set(), limit=10)
    assert len(found) == 2 and found[0].price


# ── формат «1С-иерархия» (прайс из почты 1С): шапка Номенклатура/Цена, разделы 1.1.х ──
def _xlsx_1c_hier(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Прайс-лист на 3 июля 2026 г."] + [""] * 16)
    ws.append([""] * 17)
    r = [""] * 17
    r[0], r[4], r[14] = "Артикул", "Номенклатура", "от 100 т.руб/мес"
    ws.append(r)
    r = [""] * 17
    r[14], r[15] = "Цена", "Остаток"
    ws.append(r)
    def row(art, name, price=None):
        r = [""] * 17
        r[0], r[4] = art, name
        if price:
            r[14] = price
        ws.append(r)
    row("Ту-00000002", "1. Бытовая техника")
    row("Ту-00000008", "1.1.1 Холодильники капельные")
    row("Ту-00000406", "Холодильник HOMELINE RDF-260", "17\xa0950,00")
    row("Ту-00002685", "1.1.2.3 Stinol")
    row("Ту-00002062", "Холодильник Stinol STS 185", "22\xa0010,00")
    p = tmp_path / "1c_hier.xlsx"
    wb.save(p)
    return p


def test_parse_1c_hierarchy(tmp_path):
    items = parse_price_xlsx(_xlsx_1c_hier(tmp_path))
    assert len(items) == 2
    assert items[0].name == "Холодильник HOMELINE RDF-260"
    assert items[0].price == 17950
    assert items[0].section == "Холодильники капельные"     # без «1.1.1»
    assert items[1].section == "Stinol"
    assert items[1].price == 22010
    assert items[1].article == "Ту-00002062"


def test_1c_hierarchy_search(tmp_path):
    items = parse_price_xlsx(_xlsx_1c_hier(tmp_path))
    got = select_from_price(items, "холодильники", {}, 5, taken=set())
    assert len(got) == 2


# ── построчное сопоставление списка конкретных моделей (визард /task) ─────────
def test_match_model_lines_confident_exact_copy(tmp_path):
    from content_factory.ingest.excel_price import match_model_lines
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    got = match_model_lines(items, ["Стиральная машина Beko WSRE6512"], taken=set())
    assert len(got) == 1
    assert got[0].item is not None
    assert got[0].item.name == "Стиральная машина Beko WSRE6512"
    assert got[0].candidates == []


def test_match_model_lines_partial_goes_to_candidates(tmp_path):
    from content_factory.ingest.excel_price import match_model_lines
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    # лишнее слово «инвертор» отсутствует в наименовании — не хватает для уверенного матча
    got = match_model_lines(items, ["Стиральная машина Beko WSRE6512 инвертор"], taken=set())
    assert got[0].item is None
    assert len(got[0].candidates) >= 1
    assert got[0].candidates[0].name == "Стиральная машина Beko WSRE6512"


def test_match_model_lines_no_overlap_no_candidates(tmp_path):
    from content_factory.ingest.excel_price import match_model_lines
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    got = match_model_lines(items, ["Совершенно другой несуществующий товар xyz"], taken=set())
    assert got[0].item is None
    assert got[0].candidates == []


def test_match_model_lines_respects_taken(tmp_path):
    from content_factory.ingest.excel_price import match_model_lines, item_key
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    beko = next(i for i in items if "WSRE6512" in i.name)
    got = match_model_lines(items, ["Стиральная машина Beko WSRE6512"],
                            taken={item_key(beko)})
    assert got[0].item is None                    # уже занят — не предлагаем повторно


def test_match_model_lines_preserves_order_and_skips_blank_lines(tmp_path):
    from content_factory.ingest.excel_price import match_model_lines
    items = parse_price_xlsx(_xlsx(tmp_path, ROWS))
    lines = ["Стиральная машина Candy CS4", "", "   ", "Холодильник Beko B1RCSK362S"]
    got = match_model_lines(items, lines, taken=set())
    assert len(got) == 2                           # пустые строки пропущены
    assert got[0].item.name == "Стиральная машина Candy CS4"
    assert got[1].item.name == "Холодильник Beko B1RCSK362S"


# ── два слота прайсов: свой (manual, приоритет) + почтовый (mail) ─────────────
# Грабля 2026-07-03: почта (cf-mail каждые 30 мин) молча перезаписывала
# единственный latest.xlsx поверх прайса, загруженного владельцем вручную —
# его собственные позиции переставали находиться. Теперь — раздельные слоты.
def test_load_price_slots_both_present(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "manual.xlsx")   # раздел + 2 позиции
    _xlsx(tmp_path, ROWS[3:6]).rename(tmp_path / "mail.xlsx")     # раздел + 2 позиции
    slots = load_price_slots(tmp_path)
    assert [label for label, _ in slots] == ["manual", "mail"]
    assert len(slots[0][1]) == 2 and len(slots[1][1]) == 2


def test_load_price_slots_missing_are_skipped(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "mail.xlsx")   # только почтовый
    slots = load_price_slots(tmp_path)
    assert [label for label, _ in slots] == ["mail"]


def test_load_price_slots_empty_dir(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots
    assert load_price_slots(tmp_path) == []


# ── динамические источники с наценкой (кнопка «Добавить источник», 2026-07-07) ──
def _xlsx_bytes(rows):
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№", "Артикул", "Бренд", "Наименование", "Цена (руб.)", "Заказ (шт.)"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_markup_applied_per_slot_on_load(tmp_path):
    from content_factory.ingest.excel_price import set_markup, get_markups, load_price_slots
    (tmp_path / "manual__ivanov.xlsx").write_bytes(_xlsx_bytes(
        [["Холодильники", "", "", "", "", ""],
         ["1", "10", "Beko", "Холодильник Beko X100", "10000", ""]]))
    set_markup(tmp_path, "manual__ivanov", 5)
    assert get_markups(tmp_path) == {"manual__ivanov": 5}
    items = dict(load_price_slots(tmp_path))["manual__ivanov"]
    assert items[0].price == 10500                     # +5% применены при чтении


def test_negative_markup_is_discount(tmp_path):
    from content_factory.ingest.excel_price import set_markup, load_price_slots
    (tmp_path / "manual__sale.xlsx").write_bytes(_xlsx_bytes(
        [["Чайники", "", "", "", "", ""],
         ["1", "10", "Vitek", "Чайник Vitek V1", "1000", ""]]))
    set_markup(tmp_path, "manual__sale", -7)
    items = dict(load_price_slots(tmp_path))["manual__sale"]
    assert items[0].price == 930                       # минус 7% = скидка


def test_markup_change_and_zero_removes(tmp_path):
    from content_factory.ingest.excel_price import set_markup, get_markups, load_price_slots
    (tmp_path / "manual__x.xlsx").write_bytes(_xlsx_bytes(
        [["Чайники", "", "", "", "", ""],
         ["1", "10", "Vitek", "Чайник Vitek V1", "1000", ""]]))
    set_markup(tmp_path, "manual__x", 10)
    assert dict(load_price_slots(tmp_path))["manual__x"][0].price == 1100
    set_markup(tmp_path, "manual__x", 0)               # 0 = без наценки, запись удаляется
    assert get_markups(tmp_path) == {}
    assert dict(load_price_slots(tmp_path))["manual__x"][0].price == 1000


def test_slots_without_markup_untouched(tmp_path):
    from content_factory.ingest.excel_price import set_markup, load_price_slots
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "manual.xlsx")
    (tmp_path / "manual__other.xlsx").write_bytes(_xlsx_bytes(
        [["Чайники", "", "", "", "", ""],
         ["1", "10", "Vitek", "Чайник Vitek V1", "1000", ""]]))
    set_markup(tmp_path, "manual__other", 50)
    slots = dict(load_price_slots(tmp_path))
    assert slots["manual"][0].price == 40451           # соседний слот не задет


# ── третий слот: канал-поставщик (авто-забор daily-прайса из Telegram-канала) ──
def test_load_price_slots_includes_channel_between_manual_and_mail(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "manual.xlsx")
    _xlsx(tmp_path, ROWS[6:8]).rename(tmp_path / "channel.xlsx")
    _xlsx(tmp_path, ROWS[3:6]).rename(tmp_path / "mail.xlsx")
    slots = load_price_slots(tmp_path)
    assert [label for label, _ in slots] == ["manual", "channel", "mail"]
    assert len(slots[1][1]) == 1                    # раздел Stinol + 1 позиция


def test_load_price_slots_only_channel(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "channel.xlsx")
    slots = load_price_slots(tmp_path)
    assert [label for label, _ in slots] == ["channel"]


# ── несколько ручных прайсов поставщиков (Аксёнов + БытТехОпт) одновременно ────
# Грабля 2026-07-05: один слот manual.xlsx → загрузка БытТехОпта затёрла Аксёнова.
# Теперь ручной прайс кладётся в свой слот manual__<из-имени>, грузятся все.
def test_manual_slot_name_deterministic_and_distinct():
    from content_factory.ingest.excel_price import manual_slot_name
    a = manual_slot_name("Прайс ИП Аксёнов 16.06.xlsx")
    b = manual_slot_name("БытТехОпт_20260704.xlsx")
    assert a.startswith("manual__") and b.startswith("manual__") and a != b
    assert manual_slot_name("Прайс ИП Аксёнов 16.06.xlsx") == a          # детерминизм
    assert manual_slot_name("прайс  ип  аксёнов  16.06.XLSX") == a       # регистр/пробелы
    import unicodedata                                                    # NFD (разложенные
    nfd = unicodedata.normalize("NFD", "Прайс ИП Аксёнов 16.06.xlsx")     # й/ё) == NFC:
    assert manual_slot_name(nfd) == a                                     # повторная загрузка
    assert "__" in a and "_с_" not in a and "аксёнов" in a               # тот же слот


def test_load_price_slots_multiple_manual_suppliers(tmp_path):
    from content_factory.ingest.excel_price import load_price_slots, manual_slot_name
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / f"{manual_slot_name('aksenov.xlsx')}.xlsx")
    _xlsx(tmp_path, ROWS[3:6]).rename(tmp_path / f"{manual_slot_name('byttehopt.xlsx')}.xlsx")
    slots = load_price_slots(tmp_path)
    labels = [label for label, _ in slots]
    assert len(labels) == 2 and all(label.startswith("manual__") for label in labels)
    assert sum(len(items) for _, items in slots) == 4          # обе прайса загружены


# ── top_sections: ВСЕ разделы прайсов (кнопки категорий /task) ────────────────
# Жалоба владельца 2026-07-07: показывались только топ-8 разделов — остальные
# группы товаров не выбрать кнопкой.
def test_top_sections_returns_all_sections(tmp_path):
    from content_factory.ingest.excel_price import top_sections, manual_slot_name
    rows = [["Раздел A", "", "", "", "", ""],
            ["1", "1", "X", "Товар A1", "100", ""],
            ["2", "2", "X", "Товар A2", "100", ""]]
    for i in range(12):                       # 12 разделов по 1 позиции
        rows += [[f"Раздел {i:02d}", "", "", "", "", ""],
                 [str(10 + i), str(10 + i), "Y", f"Товар {i}", "200", ""]]
    _xlsx(tmp_path, rows).rename(tmp_path / f"{manual_slot_name('p.xlsx')}.xlsx")
    secs = top_sections(tmp_path)
    assert len(secs) == 13                    # ВСЕ разделы, не топ-8
    assert secs[0] == "Раздел A"              # крупнейший — первым


def test_top_sections_optional_limit(tmp_path):
    from content_factory.ingest.excel_price import top_sections
    _xlsx(tmp_path, ROWS).rename(tmp_path / "manual.xlsx")
    assert len(top_sections(tmp_path, n=2)) == 2


def test_load_price_slots_manual_and_supplier_slots_together(tmp_path):
    # legacy manual.xlsx + новый manual__ слот сосуществуют (переходный период)
    from content_factory.ingest.excel_price import load_price_slots, manual_slot_name
    _xlsx(tmp_path, ROWS[0:3]).rename(tmp_path / "manual.xlsx")
    _xlsx(tmp_path, ROWS[3:6]).rename(tmp_path / f"{manual_slot_name('aksenov.xlsx')}.xlsx")
    slots = load_price_slots(tmp_path)
    assert len(slots) == 2 and sum(len(items) for _, items in slots) == 4


# ── словарь синонимов/транслита для поиска (/find, /make) ─────────────────────
_ALIAS_ROWS = [
    ["Стиральные машины", "", "", "", "", ""],
    ["1", "10", "Bosch", "Стиральная машина Bosch WQB245", "41990", ""],
    ["2", "11", "Candy", "Стиральная машина Candy CS4", "19990", ""],
    ["Микроволновые печи", "", "", "", "", ""],
    ["3", "12", "Samsung", "Микроволновая печь Samsung ME81", "8990", ""],
]


def test_search_alias_synonym(tmp_path):
    from content_factory.ingest.excel_price import search_items
    items = parse_price_xlsx(_xlsx(tmp_path, _ALIAS_ROWS))
    assert search_items(items, "стиралка", set()) == []            # без словаря — не находит
    got = search_items(items, "стиралка", set(), aliases={"стиралк": ["стиральная машина"]})
    assert len(got) == 2 and all("Стиральная" in i.name for i in got)


def test_search_alias_brand_translit(tmp_path):
    from content_factory.ingest.excel_price import search_items
    items = parse_price_xlsx(_xlsx(tmp_path, _ALIAS_ROWS))
    assert search_items(items, "бош", set()) == []                 # кириллица ≠ Bosch
    got = search_items(items, "бош", set(), aliases={"бош": ["bosch"]})
    assert len(got) == 1 and "Bosch" in got[0].name


def test_search_alias_combined_synonym_and_brand(tmp_path):
    from content_factory.ingest.excel_price import search_items
    items = parse_price_xlsx(_xlsx(tmp_path, _ALIAS_ROWS))
    aliases = {"стиралк": ["стиральная машина"], "бош": ["bosch"]}
    got = search_items(items, "стиралка бош", set(), aliases=aliases)
    assert len(got) == 1 and "Bosch" in got[0].name               # синоним + бренд вместе


def test_load_search_aliases_from_yaml_stems_keys(tmp_path):
    from content_factory.ingest.excel_price import load_search_aliases, stem
    y = tmp_path / "a.yaml"
    y.write_text("стиралка: [стиральная машина]\nбош: [bosch]\nсамсунг: samsung\n",
                 encoding="utf-8")
    al = load_search_aliases(y)
    assert al[stem("стиралка")] == ["стиральная машина"]           # ключи в стем-форме
    assert al[stem("бош")] == ["bosch"]
    assert al[stem("самсунг")] == ["samsung"]                      # скаляр → список


def test_load_search_aliases_missing_file_is_empty(tmp_path):
    from content_factory.ingest.excel_price import load_search_aliases
    assert load_search_aliases(tmp_path / "nope.yaml") == {}
