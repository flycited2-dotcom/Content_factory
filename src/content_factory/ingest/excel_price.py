"""Excel-прайс владельца (БытТехОпт и подобные): лист с колонками
«№ | Артикул | Бренд | Наименование | Цена | Заказ», разделы — строки, где заполнена
только первая ячейка. Источник «excel» контент-завода: выбор позиций под команду
/make N <категория> [бренд=K …] и ключи анти-дубля excel|бренд|модель."""
from __future__ import annotations
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path


@dataclass
class PriceItem:
    section: str          # раздел прайса («Холодильники с нижней морозильной камерой»)
    article: str
    brand: str
    name: str             # полное наименование позиции
    price: int


_PAREN_RE = re.compile(r"\([^)]*\)")


def _price_int(val) -> int | None:
    """'37 520,00 RUB' / '40451' / 40451.0 → int (иначе None)."""
    digits = re.sub(r"[^0-9,.]", "", str(val if val is not None else ""))
    if not digits:
        return None
    try:
        p = int(round(float(digits.replace(" ", "").replace(",", "."))))
    except ValueError:
        return None
    return p if p > 0 else None


def _parse_table(ws) -> list[PriceItem]:
    """Формат-таблица (БытТехОпт): № | Артикул | Бренд | Наименование | Цена."""
    items, section = [], ""
    for row in ws.iter_rows(values_only=True):
        a, art, brand, name, price = (list(row) + [None] * 5)[:5]
        if a and not art and not brand and not name:
            section = str(a).strip()                  # строка-раздел
            continue
        if not name or not brand:
            continue
        p = _price_int(price)
        if p is None:
            continue                                  # без цены не публикуем
        items.append(PriceItem(section=section, article=str(art or "").strip(),
                               brand=str(brand).strip(), name=str(name).strip(), price=p))
    return items


_SECTION_NUM_RE = re.compile(r"^\d+(\.\d+)*\.?\s+")


def _parse_generic(ws) -> list[PriceItem]:
    """Формат «1С-иерархия» (выгрузка прайса из 1С письмом): автопоиск шапки
    («Номенклатура/Наименование» + «Цена» в той же или следующей строке),
    разделы — строки с номенклатурой без цены (нумерация «1.1.2.3 …» срезается)."""
    rows = list(ws.iter_rows(values_only=True))
    name_col = price_col = art_col = None
    start = 0
    for i, row in enumerate(rows[:12]):
        for j, c in enumerate(row):
            t = str(c or "").strip().lower()
            if t in ("номенклатура", "наименование", "товар"):
                name_col = j
            elif "артикул" in t:
                art_col = j
        if name_col is None:
            continue
        for k in (i, i + 1):                       # «Цена» бывает строкой ниже шапки
            if k < len(rows):
                for j, c in enumerate(rows[k]):
                    if str(c or "").strip().lower() == "цена" or \
                            "цена" in str(c or "").strip().lower()[:5]:
                        price_col = j
        start = i + 2
        break
    if name_col is None or price_col is None:
        return []
    items, section = [], ""
    for row in rows[start:]:
        name = str(row[name_col] or "").strip() if len(row) > name_col else ""
        if not name:
            continue
        price = _price_int(row[price_col]) if len(row) > price_col else None
        if price:
            art = str(row[art_col] or "").strip() if art_col is not None else ""
            items.append(PriceItem(section=section, article=art, brand="",
                                   name=re.sub(r"\s+", " ", name), price=price))
        else:
            section = _SECTION_NUM_RE.sub("", name).strip()   # раздел без «1.1.2.3»
    return items


_1C_CODE_RE = re.compile(r"^(УТ-|00-)\S+")


def _parse_1c_blocks(ws) -> list[PriceItem]:
    """Формат «1С-карточки» (прайсы вида ИП Аксёнов): товар — блок строк
    (название → [характеристики] → «Код|Артикул|Цена» → значения «…RUB»),
    разделы — одиночный текст в первых колонках, колонки бренда нет."""
    items, section, name = [], "", None
    prev_blank = True
    for row in ws.iter_rows(values_only=True):
        cells = {j: str(c).strip() for j, c in enumerate(row)
                 if c is not None and str(c).strip()}
        texts = [t for t in cells.values() if "ФОТОГРАФИИ" not in t]
        if not texts:
            prev_blank = True
            continue
        if set(texts) >= {"Код", "Цена"}:             # строка-шапка блока
            prev_blank = False
            continue
        if _1C_CODE_RE.match(texts[0]):               # строка значений: код, артикул, цена
            price = next((_price_int(t) for t in texts if "RUB" in t or "руб" in t.lower()),
                         None) or (_price_int(texts[-1]) if len(texts) >= 2 else None)
            art = texts[1] if len(texts) >= 3 else ""
            if name and price:
                items.append(PriceItem(section=section, article=art, brand="",
                                       name=name, price=price))
            name, prev_blank = None, False
            continue
        low_keys = [j for j in cells if j <= 2]
        if low_keys and len(texts) == 1 and cells.get(low_keys[0]) == texts[0]:
            section = texts[0]                        # раздел/подраздел
            prev_blank = True
            continue
        if prev_blank and texts:                      # первая строка блока = название
            name = re.sub(r"\s+", " ", texts[0])
        prev_blank = False
    return items


def parse_price_xlsx(path) -> list[PriceItem]:
    """Позиции прайса. Стратегии: таблица (БытТехОпт) → блоки 1С (ИП Аксёнов и т.п.)."""
    import openpyxl                                   # тяжёлый импорт — только по нужде
    wb = openpyxl.load_workbook(Path(path), read_only=True)
    items: list[PriceItem] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        got = _parse_table(ws) or _parse_generic(ws) or _parse_1c_blocks(ws)
        items.extend(got)
    wb.close()
    return items


_SLUG_RE = re.compile(r"[^a-z0-9а-яё]+")


def manual_slot_name(filename: str) -> str:
    """Имя слота ручного прайса из имени файла: «manual__<slug>». Несколько
    прайсов поставщиков живут рядом (грабля Аксёнов/БытТехОпт 2026-07-05: один
    слот manual.xlsx → новая загрузка затирала предыдущего поставщика). Повторная
    загрузка того же файла обновляет только свой слот; регистр/пробелы не важны.
    NFC-нормализация — чтобы разложенные й/ё (NFD) из имени файла давали тот же
    слот, что и precomposed (иначе повторная загрузка задваивала бы слот)."""
    stem = unicodedata.normalize("NFC", Path(filename or "price").stem).lower()
    slug = _SLUG_RE.sub("_", stem).strip("_")[:48] or "price"
    return f"manual__{slug}"


def load_price_slots(prices_dir) -> list[tuple[str, list[PriceItem]]]:
    """Активные прайсы: ручные прайсы поставщиков «manual__*.xlsx» (все, приоритет)
    → legacy «manual.xlsx» → авто-забор из канала «channel.xlsx» → почта «mail.xlsx».
    Раздельные слоты — иначе один прайс молча перезаписывал бы другой (почта каждые
    30 мин; ручные загрузки разных поставщиков)."""
    out = []
    pdir = Path(prices_dir)
    for p in sorted(pdir.glob("manual__*.xlsx")):        # прайсы поставщиков (несколько)
        out.append((p.stem, parse_price_xlsx(p)))
    for label in ("manual", "channel", "mail"):
        p = pdir / f"{label}.xlsx"
        if p.exists():
            out.append((label, parse_price_xlsx(p)))
    return out


def extract_model(name: str, brand: str) -> str:
    """Модель = часть наименования после бренда, без скобок:
    «Холодильник Stinol STS 167 (167*60*62)» → «STS 167»."""
    clean = _PAREN_RE.sub("", name or "").strip()
    m = re.search(re.escape(brand or ""), clean, re.IGNORECASE) if brand else None
    if m:
        tail = clean[m.end():].strip(" -–—·")
        if tail:
            return re.sub(r"\s+", " ", tail)
    return re.sub(r"\s+", " ", clean)


def item_key(item: PriceItem) -> str:
    """Ключ анти-дубля (как у пилота): excel|<бренд>|<модель> (lower).
    Прайсы без колонки бренда → excel|<наименование без скобок> (lower)."""
    if item.brand.strip():
        return (f"excel|{item.brand.strip().lower()}"
                f"|{extract_model(item.name, item.brand).lower()}")
    clean = re.sub(r"\s+", " ", _PAREN_RE.sub("", item.name)).strip().lower()
    return f"excel|{clean}"


_ENDINGS_RE = re.compile(r"(иями|ями|ами|иях|ях|ах|ов|ев|ей|ий|ый|ая|яя|ое|ее|ые|ие"
                         r"|и|ы|а|я|е|о|у|ю|ь)$")


def stem(word: str) -> str:
    """Грубая основа русского слова: «генераторы/генератора/генераторов» → «генератор».
    Короткие слова (<4 после среза) не обрезаем."""
    w = (word or "").strip().lower()
    s = _ENDINGS_RE.sub("", w)
    return s if len(s) >= 4 else w


def _word_matches(word: str, text: str, aliases: dict) -> bool:
    """Слово фразы найдено в тексте: по своему стему ИЛИ по алиасу (синоним/транслит).
    Алиас-значение — фраза: срабатывает, если ВСЕ её стемы есть в тексте (напр.
    «стиралка» → «стиральная машина»: нужны и «стиральн», и «машин»)."""
    st = stem(word)
    if st in text:
        return True
    for alias in aliases.get(st, ()):
        astems = [stem(a) for a in alias.split()]
        if astems and all(a in text for a in astems):
            return True
    return False


def match_phrase(item: PriceItem, phrase: str, aliases: dict | None = None) -> int:
    """0 — не подходит; 1 — все слова фразы в разделе; 2 — все в наименовании
    (приоритетнее: «генератор» в имени ≠ АВР из раздела «Генераторы»).
    aliases — словарь синонимов/транслита (см. load_search_aliases); None = без него."""
    aliases = aliases or {}
    words = [w for w in (phrase or "").split() if w.strip()]
    if not words:
        return 0
    name = item.name.lower()
    if all(_word_matches(w, name, aliases) for w in words):
        return 2
    sec = f"{item.section} {item.name}".lower()
    if all(_word_matches(w, sec, aliases) for w in words):
        return 1
    return 0


@dataclass
class LineMatch:
    line: str
    item: PriceItem | None
    candidates: list[PriceItem]


def _line_score(item: PriceItem, line: str) -> float:
    """Доля стемов строки, найденных в наименовании (0..1)."""
    stems = [stem(w) for w in (line or "").split() if w.strip()]
    if not stems:
        return 0.0
    name = item.name.lower()
    return sum(1 for s in stems if s in name) / len(stems)


def match_model_lines(items: list[PriceItem], lines: list[str],
                      taken: set) -> list[LineMatch]:
    """Построчное сопоставление КОНКРЕТНЫХ моделей (визард /task): в отличие от
    select_from_price/search_items (одна фраза → категория/список кандидатов), тут
    каждая строка — отдельный товар и ищется отдельно. Уверенный матч (в наименовании
    найдены ВСЕ слова строки) — берём сразу; иначе не угадываем — топ-3 кандидата на
    решение владельца. Пустые строки пропускаются, порядок непустых сохраняется."""
    pool = [it for it in items if item_key(it) not in taken]
    out: list[LineMatch] = []
    for raw in lines:
        line = (raw or "").strip()
        if not line:
            continue
        scored = sorted(((s, it) for it in pool if (s := _line_score(it, line)) > 0),
                        key=lambda t: -t[0])
        if scored and scored[0][0] >= 1.0:
            out.append(LineMatch(line=line, item=scored[0][1], candidates=[]))
        else:
            out.append(LineMatch(line=line, item=None,
                                 candidates=[it for _, it in scored[:3]]))
    return out


def load_search_aliases(path) -> dict[str, list[str]]:
    """Словарь синонимов/транслита для поиска из YAML:
    «<как пишет пользователь>: [<что искать в наименовании>, …]» — напр.
    «стиралка: [стиральная машина]», «бош: [bosch]». Ключи нормализуем в стем-форму
    (ловим падежи запроса). Файла нет → пустой словарь (поиск как раньше)."""
    p = Path(path)
    if not p.exists():
        return {}
    import yaml
    raw = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    out: dict[str, list[str]] = {}
    for key, vals in raw.items():
        vlist = vals if isinstance(vals, list) else [vals]
        out.setdefault(stem(str(key)), []).extend(str(v).lower() for v in vlist)
    return out


def search_items(items: list[PriceItem], phrase: str, taken: set,
                 limit: int = 20, aliases: dict | None = None) -> list[PriceItem]:
    """Поиск позиций по фразе (без падежей + синонимы/транслит), имя-матчи первыми,
    дубли исключены. aliases — см. load_search_aliases (None = без словаря)."""
    scored = [(match_phrase(i, phrase, aliases), i) for i in items]
    pool = [i for score, i in sorted(
        [(s, i) for s, i in scored if s and item_key(i) not in taken],
        key=lambda t: -t[0])]
    return pool[:limit]


def select_from_price(items: list[PriceItem], category_kw: str, quotas: dict,
                      count: int, taken: set, aliases: dict | None = None) -> list[PriceItem]:
    """До `count` позиций категории (фраза ищется без падежей + синонимы/транслит в
    разделе и наименовании). quotas: {'beko': 3, 'stinol': 2, '*': None} — сначала
    явные квоты брендов, затем добор любыми (если задан '*' или квот нет).
    Анти-дубль по taken-ключам."""
    pool = search_items(items, category_kw, taken, limit=10 ** 9, aliases=aliases)
    out: list[PriceItem] = []

    explicit = {b: n for b, n in (quotas or {}).items() if b != "*" and n}
    for brand_kw, n in explicit.items():
        got = [i for i in pool
               if brand_kw in f"{i.brand} {i.name}".lower() and i not in out][:n]
        out.extend(got)

    fill_any = "*" in (quotas or {}) or not explicit
    if fill_any:
        for i in pool:
            if len(out) >= count:
                break
            if i not in out:
                out.append(i)
    return out[:count]
